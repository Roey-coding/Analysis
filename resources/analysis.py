from datetime import datetime, timedelta
from openpyxl import Workbook
from os import path
import openpyxl
import re

link = "Empty link"
name = "Empty name"

while(True):
	archive = input("Please enter an archive object to put the links in: ")

	if(archive == ""):
		exit()
	try:
		file = open(archive, "a+")
		break
	except:
		print("File is not valid....")

archive = archive.split('.')[0] + '.xlsx'

if path.exists(archive):
	book = openpyxl.load_workbook(archive)
	sheet = book['Sheet']
else:
	book = Workbook()
	sheet = book.active
	sheet.column_dimensions['A'].width = 20
	sheet.column_dimensions['B'].width = 25
	sheet.column_dimensions['C'].width = 250
	sheet.append(('Time', 'Name', 'Link'))
	book.save(archive)

if path.exists('root_book.xlsx'):
	root_book = openpyxl.load_workbook("root_book.xlsx")
	root_sheet = root_book['Sheet']
else:
	root_book = Workbook()
	root_sheet = root_book.active
	root_sheet.column_dimensions['A'].width = 20
	root_sheet.column_dimensions['B'].width = 25
	root_sheet.column_dimensions['C'].width = 250
	root_sheet.append(('Time', 'Name', 'Link'))
	
	root_sheet.protection.sheet = True
	root_sheet.protection.password = input("Creating a root chart....\nplease enter password for the root file: ")
	root_sheet.protection.enable()
	root_book.save('root_book.xlsx')
	

while(True):

	link = input("Please enter a link: ")
	
	if(link == ""):
		print("\nEnter pressed goodbye")
		file.close()
		exit()
	try:
		name = re.search("(https|http)\:\/\/(www\.)?(.*?)(\/|$)", link)[3]
		time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
		file.write(time + "\nName: " + name + "\nLink: " + link + "\n\n")
		file.flush()
		
		sheet.append((time, name,link))
		book.save(archive)
		
		root_sheet.append((time,name,link))
		root_book.save('root_book.xlsx')
	except:
		print("The enetered link is not valid")
