import openpyxl
import PySimpleGUI as sg
import pyautogui

resolution = pyautogui.size()

width = resolution[0]
height = resolution[1]

book = openpyxl.load_workbook("book.xlsx")
sheet = book['Sheet1']

max_row = sheet.max_row
max_column = sheet.max_column

table = []
row = []
layout = []
layout_row = []

for i in range(1, 100):
	layout_row = []
	
	layout_row.append(sg.InputText("", size = (int(width / 22 * 0.75), 2), visible = False))
	layout_row.append(sg.InputText("", size = (int(width / 22 * 0.25),2), visible = False))
	
	layout.append(layout_row)

for i in range(1, max_row + 1):
	row = []
		
	row.append(sheet.cell(row = i, column = 1).value)
	layout_row.append(sg.InputText(sheet.cell(row = i, column = 1).value, size = (int(width / 22 * 0.75), 2)))
	
	row.append(sheet.cell(row = i, column = 2).value)
	layout_row.append(sg.InputText(sheet.cell(row = i, column = 2).value, size =  (int(width / 22 * 0.25), 2)))
	
	table.append(row)
	layout[i] = [layout_row[0], layout_row[1]]
	
form = sg.Window('analysis').Layout([[sg.Column(layout, size=(width / 3, height / 2), scrollable=True, key = "Column")], 
									[sg.OK(), sg.Button('Up', key = "up"), sg.Button('Down', key = "down"), sg.Button('New row', key = "nrow")], ])


while True:	
	event, values = form.read()
	
	if event == sg.WIN_CLOSED:
		break
	elif event == "down":
		form['Column'].Widget.canvas.yview_moveto(1.0)
	elif event == "up":
		form['Column'].Widget.canvas.yview_moveto(0.0)
	elif event == "New row":
		layout.append([sg.InputText("", size = width / 21 * 0.25), sg.InputText("", size = width / 21 * 0.75)])
		form['Column'].update(layout, size=(width / 3, height / 2), scrollable=True, key = "Column")
		
form.close()